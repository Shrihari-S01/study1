"""Excel report generation."""

import asyncio
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.models.auction import Auction


class ExcelGenerator:
    """Generate Excel output for an auction notice."""

    HEADERS = [
        "Listing ID",
        "Bank Name",
        "Borrower Name",
        "Loan Number",
        "Property Type",
        "Asset Category",
        "Auction Type",
        "Movable/Immovable",
        "Reserve Price",
        "EMD",
        "Auction Date",
        "District",
        "State",
        "Contact Person",
        "Contact Number",
        "Website",
        "Confidence",
    ]

    def __init__(self) -> None:
        self.settings = get_settings()

    async def generate(self, auction: Auction) -> Path:
        """Generate an .xlsx file for the auction."""
        self.settings.excel_output_dir.mkdir(parents=True, exist_ok=True)
        path = self.settings.excel_output_dir / f"{auction.listing_id}.xlsx"
        await asyncio.to_thread(self._generate_sync, auction, path)
        return path

    def _generate_sync(self, auction: Auction, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Auction Notice"
        sheet.append(self.HEADERS)
        sheet.append(
            [
                auction.listing_id,
                auction.bank_name,
                auction.borrower_name,
                auction.loan_number,
                auction.property_type,
                auction.asset_category,
                auction.auction_type,
                auction.movable_immovable,
                auction.reserve_price,
                auction.emd,
                auction.auction_date,
                auction.district,
                auction.state,
                auction.contact_person,
                auction.contact_number,
                auction.website,
                auction.confidence_score,
            ]
        )

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(length + 2, 14), 45)
        workbook.save(path)

